import openpyxl
import json
import uuid
import re

class counter(object):
	def __init__(self):
		self.counter_int = 0

	def get_int_counter(self):
		self.counter_int+=1
		return self.counter_int-1

class alphas(object):
	def __init__(self,_name,_description):
		self.parentAlphaId = None
		self.name = _name
		self.description = _description
		self.id = str(uuid.uuid4())

class states(object):
	def __init__(self,_specialId,_name,_description,_alphaId,counter_class):
		self.id = str(uuid.uuid4())
		self.specialId = _specialId
		self.name = _name
		self.description = _description
		self.alphaId = _alphaId
		self.order = counter_class.get_int_counter()

class checkpoints(object):
	def __init__(self,_name,_description,_order,_specialId,_detailId):
		self.name = _name
		self.description = _description
		self.order = _order
		self.id = str(uuid.uuid4())
		self.specialId = _specialId
		self.detailId = _detailId

class degreesOfEvidence(object):
	def __init__(self,_iCheckableId,_checkpointId,_degreeOfEvidenceEnumValue):
		dict_degreeOfEvidenceEnumValue = {'Weak':0,'Medium':1,'Strong':2}
		self.id = str(uuid.uuid4())
		self.iCheckableId = _iCheckableId
		self.typeOfEvidence = True
		self.checkpointId = _checkpointId
		self.degreeOfEvidenceEnumValue = dict_degreeOfEvidenceEnumValue[_degreeOfEvidenceEnumValue]

mas_alphas = []
mas_states = []
mas_checkpoints = []
mas_degreesOfEvidence = []

def main():
	wb = openpyxl.load_workbook("Protocol.xlsx",read_only=True)
	sheet = wb["Полный протокол"]
	row_mas_alpha = [2,21,41,63,86,107,126]
	counter_class = counter()
	for row_id in row_mas_alpha:
		alphaName = sheet.cell(row = row_id-1, column=1)
		print(alphaName.value)
		mas_alphas.append(
					alphas(
						alphaName.value,
						""
						)
					)
		col_id = 1
		while(col_id<=sheet.max_column):

			stateId = sheet.cell(row = row_id, column=col_id)

			if(stateId.value != None):

				stateName = sheet.cell(row = row_id+1, column=col_id)
				print(stateName.value)

				mas_states.append(
					states(
						stateId.value,
						stateName.value,
						"",
						mas_alphas[len(mas_alphas)-1].id,
						counter_class
						)
					)

				col_id_checkpoint = col_id

				while(sheet.cell(row = row_id, column=col_id_checkpoint).value == stateId.value or sheet.cell(row = row_id, column=col_id_checkpoint).value == None):
					if(col_id_checkpoint > sheet.max_column):
						break
					if(sheet.cell(row = row_id+2, column=col_id_checkpoint).value != None):
						checkpoint_name = sheet.cell(row = row_id+2, column=col_id_checkpoint)
						checkpoint_desc = sheet.cell(row = row_id+3, column=col_id_checkpoint)
						checkpoint_specialID = sheet.cell(row = row_id+4, column=col_id_checkpoint)
						print(checkpoint_specialID.value)
						mas_checkpoints.append(
							checkpoints(
								checkpoint_name.value,
								checkpoint_desc.value,
								mas_states[len(mas_states)-1].order,
								checkpoint_specialID.value,
								mas_states[len(mas_states)-1].id
							)
						)

					col_id_checkpoint+=1
			col_id+=1

	for row_id in row_mas_alpha:
		alphaName = sheet.cell(row = row_id, column=1)
		print(alphaName.value)
		col_id = 1
		while(col_id<=sheet.max_column):
			stateId = sheet.cell(row = row_id, column=col_id)
			if(stateId.value != None):
				stateName = sheet.cell(row = row_id+1, column=col_id)
				col_id_checkpoint = col_id
				while(sheet.cell(row = row_id, column=col_id_checkpoint).value == stateId.value or sheet.cell(row = row_id, column=col_id_checkpoint).value == None):
					if(col_id_checkpoint > sheet.max_column):
						break
					if(sheet.cell(row = row_id+2, column=col_id_checkpoint).value != None):
						checkpoint_name = sheet.cell(row = row_id+2, column=col_id_checkpoint)
						checkpoint_id = None
						for checkpoint in mas_checkpoints:
							if checkpoint.name == checkpoint_name.value:
								checkpoint_id = checkpoint.id
								break
						row_id_deg = row_id+5
						while(sheet.cell(row = row_id_deg, column=col_id_checkpoint).value != "Manager op"):
							checkpoint_parant_name = sheet.cell(row = row_id_deg, column=col_id_checkpoint)
							print(checkpoint_parant_name.value)
							Icheckpoint_id = None
							if re.fullmatch(r'\D+\d+',str(checkpoint_parant_name.value)):
								if re.fullmatch(r'\D+\d{1}',str(checkpoint_parant_name.value)):
									for state in mas_states:
										if state.specialId == checkpoint_parant_name.value:
											Icheckpoint_id = state.id
											break
								else:
									for checkpoint in mas_checkpoints:
										if checkpoint.specialId == checkpoint_parant_name.value:
											Icheckpoint_id = checkpoint.id
											break
								degree = sheet.cell(row = row_id_deg, column= col_id_checkpoint+2)
								mas_degreesOfEvidence.append(
									degreesOfEvidence(
										Icheckpoint_id,
										checkpoint_id,
										degree.value
										)
									)
							row_id_deg+=1
					col_id_checkpoint+=1
			col_id+=1





	str_json_alphas = "["
	for alpha in mas_alphas:
		str_json_alphas+='{'
		str_json_alphas+="'parentAlphaId':"+ str(alpha.parentAlphaId).replace("'",r"\'") +','
		str_json_alphas+="'name':'"+str(alpha.name).replace("'",r"\'") +"',"
		str_json_alphas+="'description':'"+str(alpha.description).replace("'",r"\'") +"',"
		str_json_alphas+="'id':'"+str(alpha.id).replace("'",r"\'")+"'"
		str_json_alphas+='},'
	str_json_alphas= str_json_alphas[:len(str_json_alphas)-1] +']'
	f = open('json_output/alphas.json','w')
	f.write(json.dumps(eval(str_json_alphas), indent=4))
	f.close()

	str_json_state = "["
	for state in mas_states:
		str_json_state+='{'
		str_json_state+="'id':'" + str(state.id).replace("'",r"\'") +"',"
		str_json_state+="'specialId':'"+ str(state.specialId).replace("'",r"\'") +"',"
		str_json_state+="'name':'"+str(state.name).replace("'",r"\'")+"',"
		str_json_state+="'description':'"+str(state.description).replace("'",r"'")+"',"
		str_json_state+="'alphaId':'"+str(state.alphaId).replace("'",r"\'")+"',"
		str_json_state+="'order':"+str(state.order).replace("'",r"\'")
		str_json_state+='},'

	str_json_state = str_json_state[:len(str_json_state)-1] + "]"
	f = open('json_output/states.json','w')
	f.write(json.dumps(eval(str_json_state), indent=4))
	f.close()

	str_json_checkpoints = "["
	for checkpoint in mas_checkpoints:
		str_json_checkpoints+='{'
		str_json_checkpoints+="'name':'"+str(checkpoint.name).replace("'",r"\'")+"',"
		str_json_checkpoints+="'description':'"+str(checkpoint.description).replace("'",r"\'")+"',"
		str_json_checkpoints+="'order':"+str(checkpoint.order).replace("'",r"\'")+','
		str_json_checkpoints+="'id':'"+str(checkpoint.id).replace("'",r"\'")+"',"
		str_json_checkpoints+="'specialId':'"+str(checkpoint.specialId).replace("'",r"\'")+"',"
		str_json_checkpoints+="'detailId':'"+str(checkpoint.detailId).replace("'",r"\'")+"'"
		str_json_checkpoints+="},"
	str_json_checkpoints = str_json_checkpoints[:len(str_json_checkpoints)-1]+"]"
	f = open('json_output/checkpoints.json','w')
	f.write(json.dumps(eval(str_json_checkpoints), indent=4))
	f.close()

	str_json_degreesOfEvidence = "["
	for degree in mas_degreesOfEvidence:
		str_json_degreesOfEvidence+='{'
		str_json_degreesOfEvidence+="'id':'"+str(degree.id).replace("'",r"\'")+"',"
		str_json_degreesOfEvidence+="'iCheckableId':'"+str(degree.iCheckableId).replace("'",r"\'")+"',"
		str_json_degreesOfEvidence+="'typeOfEvidence':"+str(degree.typeOfEvidence)+','
		str_json_degreesOfEvidence+="'checkpointId':'"+str(degree.checkpointId)+"',"
		str_json_degreesOfEvidence+="'degreeOfEvidenceEnumValue':"+str(degree.degreeOfEvidenceEnumValue)
		str_json_degreesOfEvidence+="},"
	str_json_degreesOfEvidence = str_json_degreesOfEvidence[:len(str_json_degreesOfEvidence)-1]+"]"
	f = open('json_output/degreesOfEvidence.json','w')
	f.write(json.dumps(eval(str_json_degreesOfEvidence), indent=4))
	f.close()

if __name__ == '__main__':
	main()