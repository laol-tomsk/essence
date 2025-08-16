import json
import openpyxl
import uuid
import re


class checkpoints(object):
	def __init__(self,_name,_description,_order,_id,_specialId,_detailId):
		self.name = _name
		self.description = _description
		self.order = _order
		self.id = _id
		self.specialId = _specialId
		self.detailId = _detailId
		self.degreeOfEvidenceEnumValue = 3

def main():
	wb = openpyxl.load_workbook("Protocol.xlsx",read_only=True)
	sheet = wb["Полный протокол"]
	row_mas_alpha = [6,25,45,67,90,111,130]
	_dict = {'Weak':0,'Medium':1,'Strong':2}
	with open("n_checkpoints.json") as f:
		json_checkpoint= json.load(f)

	dict_checkpoints = {}
	mas_checkpoints = []
	dop_mas = []
	for checkpoint_d in json_checkpoint:
		checkpoint = checkpoints(checkpoint_d["name"],checkpoint_d["description"],checkpoint_d["order"],checkpoint_d["id"],checkpoint_d["specialId"],checkpoint_d["detailId"])
		if checkpoint_d["specialId"] == "None":
			dop_mas.append(checkpoint)
		dict_checkpoints.update({checkpoint_d["specialId"]:checkpoint})

	for row_id in row_mas_alpha:
		col_id = 1
		while (col_id<= sheet.max_column):
			if(sheet.cell(row = row_id, column=col_id).value == None):
				break
			checkpointName = sheet.cell(row = row_id, column=col_id).value
			print(row_id)
			row_id_castom = row_id
			while(sheet.cell(row = row_id_castom, column=col_id).value != "Manager op"):
				row_id_castom+=1

			DegreeOfEvidence = sheet.cell(row = row_id_castom, column=col_id+2).value

			dict_checkpoints[checkpointName].degreeOfEvidenceEnumValue = _dict[DegreeOfEvidence]
			mas_checkpoints.append(dict_checkpoints[checkpointName])
			col_id+=3


	str_json_checkpoints = "["
	for checkpoint in mas_checkpoints+dop_mas:
		str_json_checkpoints+='{'
		str_json_checkpoints+="'name':'"+str(checkpoint.name).replace("'",r"\'")+"',"
		str_json_checkpoints+="'description':'"+str(checkpoint.description).replace("'",r"\'")+"',"
		str_json_checkpoints+="'order':"+str(checkpoint.order).replace("'",r"\'")+','
		str_json_checkpoints+="'id':'"+str(checkpoint.id).replace("'",r"\'")+"',"
		str_json_checkpoints+="'specialId':'"+str(checkpoint.specialId).replace("'",r"\'")+"',"
		str_json_checkpoints+="'degreeOfEvidenceEnumValueManagerOpinion':"+str(checkpoint.degreeOfEvidenceEnumValue).replace("'",r"\'")+","
		str_json_checkpoints+="'detailId':'"+str(checkpoint.detailId).replace("'",r"\'")+"'"
		str_json_checkpoints+="},"
	str_json_checkpoints = str_json_checkpoints[:len(str_json_checkpoints)-1]+"]"
	f = open('checkpoints.json','w')
	f.write(json.dumps(eval(str_json_checkpoints), indent=4))
	f.close()

if __name__ == "__main__":
	main()