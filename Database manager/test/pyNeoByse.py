from py2neo import Graph, GraphService, Node, Relationship
import os
import xml.etree.ElementTree as ET
import time
import openpyxl
import re
import json

class NeoByse(object):

	#Конструктор класса
	def __init__(self,URL,Account):
		self.GeneralСonnection = GraphService(URL, auth=Account)
		self.URL = URL
		self.Account = Account

	#Метод создания подключение с базой данных по её именни
	def __congraph__(self,name):
		connect = Graph(self.URL, auth=self.Account, name = name)
		return connect

	#Метод создания транзакции для базы данных
	def __tx_database__(self,name_database):
		return self.__congraph__(name_database).begin()

	#Метод подключение к системной базе данных сервера
	def __congraphsystem__(self):
		return self.GeneralСonnection.system_graph

	#Метод подключение к дефолтной базе данных сервера
	def __congraphmain__(self):
		return self.GeneralСonnection.default_graph

	#Метод создания транзакции для изменения базы данных
	def __settransaction__(self,graph,text):
		transaction = graph.begin()
		transaction.run(text)
		graph.commit(transaction)

	#Метод создания транзакции для получения данных с сервера
	def __gettransaction__(self,graph,text):
		transaction = graph.begin()
		return transaction.run(text).data()

	#Метод проверки на состояние базы данных
	def __checkingstatus__(self,name_database,status,databases_status):
		if([name_database,status] in databases_status):
			return True
		else:
			return False

	#Метод проверки на то, что база имеется на сервере
	def __checkbaseexistence__(self,name_database,get_bases):
		if(name_database in get_bases):
			return True
		else:
			return False

	#Функция получения всех баз данных которые есть на сервере
	def get_bases(self):
		mas_name_bases = []
		for name_bases in iter(self.GeneralСonnection):
			mas_name_bases.append(name_bases)
		print("Bases received")
		return mas_name_bases
		pass

	#Функция получение состояния всех баз на сервере
	def show_databases_status(self):
		data = self.__gettransaction__(self.__congraphsystem__(),"SHOW DATABASES")
		mas_database_status = []
		for status in data:
			database_status = [status['name'],status['requestedStatus']]
			mas_database_status.append(database_status)
		print("Status received")
		return mas_database_status
		pass

	#Функция для сохдания новой ноды
	def create_new_node(self,identity,dictionary,name_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			dop_str = " {"
			for key, value in dictionary.items():
				dop_str+=str(key)+ ":'" + str(value) + "',"
			dop_str = dop_str[:len(dop_str)-1]+"})"
			command = "CREATE (:"+str(identity) + dop_str
			self.__settransaction__(self.__congraph__(name_database),command)
			print("Node created")
		else:
			print("BaseError: Database does not exist")

	#Функция для создание новой связи
	def create_new_relationship(self,nodeA,nodeB,name_datebase):
		if (self.__checkbaseexistence__(name_datebase,self.get_bases())):
			tx = self.__tx_database__(name_datebase)
			print(nodeA)
			print(nodeB)
			tx.create(Relationship(nodeA,"S",nodeB))
			tx.commit()
			print("Relationship created")
		else:
			print("BaseError: Database does not exist")
		pass

	#Функция для поиска ноды в базеданных
	def searth_node(self,name_node,name_database):
		#if (self.__checkbaseexistence__(name_database,self.get_bases())):
		tx = self.__congraph__(name_database)
		return tx.nodes.match(name=name_node).first()
		#else:
			#print("BaseError: Database does not exist")
		pass

	#Функция для получения всех нод базы данных
	def return_all_nodes(self,name_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			tx = self.__congraph__(name_database)
			return tx.nodes.match().all()
		else:
			print("BaseError: Database does not exist")
		pass

	#Функция создания базы данных
	def сreate_database(self,name_database):
		if (not self.__checkbaseexistence__(name_database,self.get_bases())):
			self.__settransaction__(self.__congraphsystem__(),"CREATE DATABASE " + name_database)
			print("Base created")
		else:
			print("BaseError: The database has already been created")
		pass

	#Функция удаления базы данных
	def drop_database(self,name_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			self.__settransaction__(self.__congraphsystem__(),"DROP DATABASE " + name_database)
			print("Base deleted")
		else:
			print("BaseError: Database does not exist")
		pass

	#Функция запуска базы данных
	def start_database(self,name_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			if(not self.__checkingstatus__(name_database,"online", self.show_databases_status())):
				self.__settransaction__(self.__congraphsystem__(),"START DATABASE " + name_database)
				print("Base launched")
			else:
				print("StatusError: The database is already running")
		else:
			print("BaseError: The base does not exist")
		pass

	#Функция остановки базы данных
	def stop_database(self,name_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			if(not self.__checkingstatus__(name_database,"offline", self.show_databases_status())):
				self.__settransaction__(self.__congraphsystem__(),"STOP DATABASE " + name_database)
				print("Base stopped")
			else:
				print("StatusError: The database is already stopped")
		else:
			print("BaseError: The base does not exist")
		pass

	#Функция отправки запроса на базу данных
	def send_request(self,request,database):
		if (self.__checkbaseexistence__(database,self.get_bases())):
			if(self.__checkingstatus__(database,"online", self.show_databases_status())):
				try:
					print("Request completed")
					if(("CREATE" in request) or ("DELETE" in request)):
						self.__settransaction__(self.__congraph__(database),request)
					elif("MATCH" in request):
						data = self.__gettransaction__(self.__congraph__(database),request)
						return data
				except:
					print("RequestError: The request returned with an error")
			else:
				print("BaseError: The database stopped")
		else:
			print("BaseError: The base does not exist")
		pass

	#Функция для создания копии базы данных
	def copy_base(self,name_database,rename,rename_database):
		if (self.__checkbaseexistence__(name_database,self.get_bases())):
			self.stop_database(name_database)
			if (rename):
				os.system("D:/Neo4j/relate-data/dbmss/dbms-f72e0abb-4380-4f53-a258-4c36cadf56bf/bin/neo4j-admin copy --from-database="+name_database +" --to-database="+rename_database)
			else:
				os.system("D:/Neo4j/relate-data/dbmss/dbms-f72e0abb-4380-4f53-a258-4c36cadf56bf/bin/neo4j-admin copy --from-database="+name_database +" --to-database="+name_database+"copy")

			if(rename):
				self.сreate_database(rename_database)
			else:
				self.start_database(name_database)
				self.сreate_database(name_database+"copy")
			print("Copy was successful")
		else:
			print("BaseError: There is no such database on the server")
		pass

	def rename_database(self,name_database,rename_database):
		if (not self.__checkbaseexistence__(rename_database,self.get_bases())):
			self.copy_base(name_database,True,rename_database)
			self.drop_database(name_database)
			print("Base renamed")
		else:
			print("BaseError: Such a database already exists on the server")
		pass

	#Удаление всех баз данных
	def all_database_remuve(self):
		mas_database = self.get_bases()
		for base in mas_database:
			if base != "neo4j" and base != "system":
				self.drop_database(base)
			elif(base == "neo4j"):
				self.send_request("MATCH (p1)-[r]->(p2) DELETE r","neo4j")
				self.send_request("MATCH (p) DELETE p","neo4j")
		print("Factory reset was successful")
		pass

	#Функция для создания базы данных из файла
	def create_database_from_file(self,name_database,PATH):
		start_time = time.time()
		if (not self.__checkbaseexistence__(name_database,self.get_bases())):
			self.сreate_database(name_database)
			parser = parser_XML(PATH)
			for identity,dictionary in parser.return_state_cpt():
				self.create_new_node(identity,dictionary,name_database)
			for identity,dictionary in parser.return_state_decision():
				self.create_new_node(identity,dictionary,name_database)
			for node in self.return_all_nodes(name_database):
				if "parents" in node.keys():
					for parent in node.get("parents").split(" "):
						self.create_new_relationship(self.searth_node(parent,name_database),self.searth_node(node.get("name"),name_database),name_database)
				else:
					print("NodeError: Node not parents")
			print("--- %s seconds ---" % (time.time() - start_time))
		else:
			print("BaseError: Database does not exist")

	def create_database_from_protocol(self,name_database,PATH,name_table):
		start_time = time.time()
		if (not self.__checkbaseexistence__(name_database,self.get_bases())):
			self.сreate_database(name_database)
			parser = parser_protocol(PATH,name_table)
			for identity,dictionary in parser.return_cpt():
				self.create_new_node(identity,dictionary,name_database)
			for identity,dictionary in parser.return_decision():
				self.create_new_node(identity,dictionary,name_database)
			for node in self.return_all_nodes(name_database):
				if "parents" in node.keys():
					for parent in node.get("parents").split(" "):
						self.create_new_relationship(self.searth_node(parent,name_database),self.searth_node(node.get("name"),name_database),name_database)
				else:
					print("NodeError: Node not parents")
			print("--- %s seconds ---" % (time.time() - start_time))
		else:
			print("BaseError: Database does not exist")

#Класс XML парсера
class parser_XML(object):
	def __init__(self,PATH):
		self.ROOT = ET.parse(PATH).getroot()[0]

	#Функция возврата поддерева
	def __return_cpt_list__(self):
		return self.ROOT.findall('cpt')
	#Функция возврата поддерева
	def __return_decision_list__(self):
		return self.ROOT.findall('decision')

	#Функция получение всех данных о ноде
	def return_state_cpt(self):

		for elem_cpt in self.__return_cpt_list__():
			name = elem_cpt.get('id')
			state_1 = elem_cpt.findall('state')[0].get('id')
			state_2 = elem_cpt.findall('state')[1].get('id')
			state = state_1 + ' ' + state_2
			parents = None
			try:
				parents = elem_cpt.find('parents').text
			except:
				pass
			probabilities = elem_cpt.find('probabilities').text
			if parents == None:
				d ={'name': name, 'state': state, 'probabilities':probabilities, 'check_parents': False}
			else:
				d={'name': name, 'state': state,'parents': parents, 'probabilities':probabilities, 'check_parents': True}

			yield 'cpt',d

	#Функция получение всех данных о ноде
	def return_state_decision(self):
		for elem_decision in self.__return_decision_list__():
			name = elem_decision.get('id')
			state_1 = elem_decision.findall('state')[0].get('id')
			state_2 = elem_decision.findall('state')[1].get('id')
			state = state_1 + ' ' + state_2
			d = {'name':name,'state': state}
			yield 'decision', d

class parser_protocol(object):
	def __init__(self,PATH,name_table):
		self.Table = openpyxl.load_workbook(PATH)[name_table]
		self.Mas_inter_row = [6,25,45,67,90,111,130]
		self.Mas_decision = []
		self.Dict_perants_dicision = {}
		self.Вegree_of_influence = {'Strong': 5, 'Medium': 2, 'Weak': 1}

	def return_cpt(self):
		for row in self.Mas_inter_row:
			for col in range(1,self.Table.max_column+1,3):
				row_n = row
				name = self.Table.cell(row = row_n, column=col).value
				if name == None:
					break
				self.Mas_decision.append(name[:-1])
				if name[:-1] in self.Dict_perants_dicision.keys():
					self.Dict_perants_dicision[name[:-1]].append(name)
				else:
					self.Dict_perants_dicision.update({name[:-1]:[name]})
				parents = ''
				degree = {}

				while(self.Table.cell(row = row_n+1, column=col).value != 'Manager op'):
					row_n +=1
					node = self.Table.cell(row = row_n, column=col).value
					#print(node)
					if node == None:
						continue
					else:
						parents += node + ' '
						degree.update({node:self.Вegree_of_influence[self.Table.cell(row = row_n, column=col+2).value]})
				parents = (name+'S ')+parents
				parents = (name+'C ')+parents
				degree2 = {name+'C':self.Вegree_of_influence[self.Table.cell(row = row_n+1, column=col+2).value],name+'S':self.Вegree_of_influence['Weak']}
				degree2.update(degree)
				yield 'decision', {'name': name+'C'}
				yield 'cpt_not', {'name': name+'S'}
				yield 'cpt', {'name': name,'parents': parents[:-1], 'degree': json.dumps(degree2)}

	def return_decision(self):
		self.Mas_decision = list(set(self.Mas_decision))
		print(self.Mas_decision)
		print(self.Dict_perants_dicision)
		input()
		for name in self.Mas_decision:
			parents = ''
			for name_p in self.Dict_perants_dicision[name]:
				parents += name_p + ' '

			yield 'cpt', {'name': name,'parents': parents[:-1]}











